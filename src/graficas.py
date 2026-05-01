# Módulo de gráficas y exportación a Excel.
# Última modificación: 29/04/2026
# Autor: Adrián Humberto Cavazos Leal

import matplotlib.pyplot as plt
import requests
import math
import json
from pathlib import Path
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_DIR_SALIDA     = Path(__file__).resolve().parent.parent / "reportes"
_HISTORIAL_JSON = Path(__file__).resolve().parent.parent / "historial_comparativa.json"
_MAX_CIUDADES   = 4
_KEY_GRAFICAS   = "64e5e1c01b176ce25da569b8627e31a9"

_DIR_SALIDA.mkdir(parents=True, exist_ok=True)


#One Call 3.0

def _obtener_onecall(ciudad: str) -> dict:
    import api_cliente
    try:
        coords = api_cliente.dato_clima(ciudad)
        lat, lon = coords["coord"]["lat"], coords["coord"]["lon"]
        url = (f"https://api.openweathermap.org/data/3.0/onecall"
               f"?lat={lat}&lon={lon}&units=metric&exclude=minutely,alerts"
               f"&appid={_KEY_GRAFICAS}")
        datos = requests.get(url).json()
        datos["_lat"] = lat
        datos["_lon"] = lon
        return datos
    except Exception:
        return {}

def _procesar_onecall(datos_oc: dict) -> tuple:
    # Toma las 24 entradas en orden cronológico desde la hora actual
    # Índice 0 = ahora, índice 1 = ahora+1h, ..., índice 23 = ahora+23h
    hourly    = datos_oc.get("hourly", [])[:24]
    tz_offset = datos_oc.get("timezone_offset", 0)

    temps_24h = [round(e["temp"], 1) for e in hourly]

    # Rellenar si faltan entradas al final
    while len(temps_24h) < 24:
        temps_24h.append(temps_24h[-1] if temps_24h else 0.0)

    # Máx, mín y media calculadas desde el mismo temps_24h que se grafica
    temp_max = max(temps_24h)
    temp_min = min(temps_24h)

    return temps_24h, temp_max, temp_min, tz_offset


#Historial JSON

def _cargar_historial() -> list:
    if _HISTORIAL_JSON.exists():
        with open(_HISTORIAL_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def _guardar_historial(historial: list) -> None:
    with open(_HISTORIAL_JSON, "w", encoding="utf-8") as f:
        json.dump(historial, f, ensure_ascii=False, indent=2)

def _registrar_ciudad(ciudad: str, temps_24h: list, temp_max, temp_min) -> list:
    historial = _cargar_historial()
    if len(historial) >= _MAX_CIUDADES:
        historial = []
    historial = [c for c in historial if c["ciudad"].lower() != ciudad.lower()]

    vals   = [t for t in temps_24h if t is not None]
    media  = round(sum(vals) / len(vals), 1) if vals else 0.0
    maxima = round(temp_max, 1) if temp_max is not None else round(max(vals), 1) if vals else 0.0
    minima = round(temp_min, 1) if temp_min is not None else round(min(vals), 1) if vals else 0.0

    historial.append({
        "ciudad": ciudad,
        "media":  media,
        "maxima": maxima,
        "minima": minima,
        "fecha":  datetime.now().strftime("%Y-%m-%d")
    })
    _guardar_historial(historial)
    return historial


#Punto de entrada

def generar_reporte(resultado_analizador: dict, datos_api: dict, ciudad: str) -> None:
    import api_cliente

    fecha    = resultado_analizador.get("fecha", datetime.now().strftime("%Y-%m-%d"))
    clima    = resultado_analizador.get("clima", "")
    alerta   = resultado_analizador.get("alerta", "")
    recomend = resultado_analizador.get("recomendacion", "")
    humedad  = datos_api.get("main", {}).get("humidity")
    viento   = datos_api.get("wind", {}).get("speed")

    datos_oc                          = _obtener_onecall(ciudad)
    temps_24h, t_max, t_min, tz_off  = _procesar_onecall(datos_oc)

    # Fallback senoidal si One Call no devuelve datos
    if not datos_oc.get("hourly"):
        temperatura = resultado_analizador.get("temperatura")
        temps_24h   = _curva_senoidal(temperatura)
        t_max, t_min = None, None

    historial = _registrar_ciudad(ciudad, temps_24h, t_max, t_min)

    lat = datos_oc.get('_lat', 'N/D')
    lon = datos_oc.get('_lon', 'N/D')
    grafica_consulta(fecha, clima, ciudad, temps_24h, tz_off)
    exportar_excel(fecha, clima, humedad, viento, alerta, recomend, ciudad, historial, lat, lon)

    if len(historial) > 1:
        grafica_comparativa(historial)


#Fallback senoidal

def _curva_senoidal(temp_actual: float) -> list:
    hora_actual = datetime.now().hour
    base  = temp_actual - 4
    temps = [
        round(base + (temp_actual - base) * math.sin(math.pi * max(h - 5, 0) / 18), 1)
        if 5 <= h <= 23 else base
        for h in range(24)
    ]
    temps[hora_actual] = temp_actual
    return temps


#Gráfica 1: temperatura del día

def grafica_consulta(
    fecha: str, clima: str, ciudad: str,
    temps_24h: list, tz_offset: int = 0
) -> None:
    tz_ciudad   = timezone(timedelta(seconds=tz_offset))
    hora_inicio = datetime.now(tz=tz_ciudad).hour
    horas       = list(range(24))

    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(horas, temps_24h, color="#E8622A", linewidth=2.5, zorder=2)
    ax.fill_between(horas, temps_24h, alpha=0.12, color="#E8622A")

    temp_punto = temps_24h[0]
    ax.scatter([0], [temp_punto], color="#C04010", s=100, zorder=5)
    ax.annotate(f"{temp_punto}°C",
                xy=(0, temp_punto), xytext=(0, 13),
                textcoords="offset points", ha="center",
                fontsize=10, fontweight="bold", color="#C04010")

    ax.set_title(f"Temperatura del día — {ciudad} ({fecha})", fontsize=13, fontweight="bold", pad=14)
    ax.set_xlabel("Hora del día", fontsize=11)
    ax.set_ylabel("Temperatura (°C)", fontsize=11)
    etiquetas   = [f"{(hora_inicio + i) % 24}h" for i in range(24)]
    ax.set_xticks(range(24))
    ax.set_xticklabels(etiquetas, rotation=45, fontsize=8)
    ax.grid(axis="y", linestyle="--", alpha=0.45)
    ax.spines[["top", "right"]].set_visible(False)
    ax.set_facecolor("#FAFAFA")

    plt.figtext(0.5, -0.02, f"Condición: {clima}", ha="center", fontsize=9, color="#555555")
    plt.tight_layout()

    nombre_img = _DIR_SALIDA / f"grafica_{ciudad.lower().replace(' ', '_')}_{fecha}.png"
    plt.savefig(nombre_img, dpi=150, bbox_inches="tight")
    plt.show()
    print(f"Gráfica guardada: {nombre_img}")


#Gráfica 2:

def grafica_comparativa(historial: list) -> None:
    ciudades = [r["ciudad"] for r in historial]
    maximas  = [r["maxima"] for r in historial]
    medias   = [r["media"]  for r in historial]
    minimas  = [r["minima"] for r in historial]

    x, ancho = range(len(ciudades)), 0.22
    fig, ax  = plt.subplots(figsize=(10, 5))

    for offset, valores, color, etiqueta in [
        (-ancho, maximas, "#D94F3D", "T. máxima °C"),
        (0,      medias,  "#4DA87B", "T. media °C"),
        (ancho,  minimas, "#3A7FBF", "T. mínima °C"),
    ]:
        ax.bar([xi + offset for xi in x], valores, width=ancho,
               color=color, label=etiqueta, alpha=0.85, zorder=2)

    ax.set_title("Comparativa de temperaturas por ciudad", fontsize=13, fontweight="bold", pad=14)
    ax.set_ylabel("Temperatura (°C)", fontsize=11)
    ax.set_xticks(list(x))
    ax.set_xticklabels(ciudades, fontsize=10)
    ax.grid(axis="y", linestyle="--", alpha=0.4, zorder=1)
    ax.spines[["top", "right"]].set_visible(False)
    ax.set_facecolor("#FAFAFA")
    ax.legend(fontsize=9, loc="upper right")
    plt.tight_layout()

    nombre_img = _DIR_SALIDA / f"comparativa_{datetime.now().strftime('%Y-%m-%d')}.png"
    plt.savefig(nombre_img, dpi=150, bbox_inches="tight")
    plt.show()
    print(f"Gráfica comparativa guardada: {nombre_img}")


#Excel

def exportar_excel(
    fecha: str, clima: str, humedad, viento,
    alerta: str, recomendacion: str, ciudad: str, historial: list,
    lat=None, lon=None
) -> str:
    nombre_archivo = _DIR_SALIDA / f"clima_{ciudad.lower().replace(' ', '_')}_{fecha}.xlsx"

    fe  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    re  = PatternFill("solid", start_color="2D6A9F")
    ac  = Alignment(horizontal="center", vertical="center")
    bo  = Border(bottom=Side(style="medium", color="1A4E7A"),
                 right=Side(style="thin", color="1A4E7A"))
    rp  = PatternFill("solid", start_color="EAF2FB")

    # Tomar valores del historial para consistencia con la gráfica
    reg       = next((r for r in historial if r["ciudad"].lower() == ciudad.lower()), {})
    temp_med  = reg.get("media",  "N/D")
    temp_max  = reg.get("maxima", "N/D")
    temp_min  = reg.get("minima", "N/D")

    wb = Workbook()

    #Hoja 1
    ws = wb.active
    ws.title = "Datos Climáticos"

    ws.merge_cells("A1:B1")
    ws["A1"].value     = f"Reporte Climático — {ciudad}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1A3A5C")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill      = PatternFill("solid", start_color="D6E8F7")
    ws.row_dimensions[1].height = 28

    for col, texto in enumerate(["Parámetro", "Valor"], start=1):
        c = ws.cell(row=2, column=col, value=texto)
        c.font = fe; c.fill = re; c.alignment = ac; c.border = bo
    ws.row_dimensions[2].height = 22

    filas = [
        ("Fecha de consulta",          fecha),
        ("Ciudad",                     ciudad),
        ("Temperatura máxima (°C)",    temp_max),
        ("Temperatura media (°C)",     temp_med),
        ("Temperatura mínima (°C)",    temp_min),
        ("Latitud",                    lat if lat is not None else "N/D"),
        ("Longitud",                   lon if lon is not None else "N/D"),
        ("Condición climática",        clima),
        ("Humedad (%)",                humedad if humedad is not None else "N/D"),
        ("Velocidad del viento (m/s)", viento  if viento  is not None else "N/D"),
        ("Alerta",                     alerta if alerta else "Sin alertas"),
        ("Recomendación",              recomendacion),
    ]

    for i, (param, valor) in enumerate(filas):
        fila = i + 3
        ws.cell(fila, 1, param).font = Font(name="Arial", size=10, bold=True)
        cv = ws.cell(fila, 2, valor)
        cv.font = Font(name="Arial", size=10)
        cv.alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2 == 0:
            ws.cell(fila, 1).fill = rp; cv.fill = rp
        ws.row_dimensions[fila].height = 18

    ws.row_dimensions[len(filas) + 2].height = 45
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55
    ws.cell(len(filas) + 4, 1,
            f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = \
        Font(name="Arial", size=9, italic=True, color="888888")

    #Hoja 2
    if len(historial) > 1:
        ws2 = wb.create_sheet("Comparativa")

        ws2.merge_cells("A1:D1")
        ws2["A1"].value     = "Comparativa de Temperaturas"
        ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color="1A3A5C")
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2["A1"].fill      = PatternFill("solid", start_color="D6E8F7")
        ws2.row_dimensions[1].height = 28

        for col, texto in enumerate(["Ciudad", "T. Máxima (°C)", "T. Media (°C)", "T. Mínima (°C)"], start=1):
            c = ws2.cell(row=2, column=col, value=texto)
            c.font = fe; c.fill = re; c.alignment = ac; c.border = bo
        ws2.row_dimensions[2].height = 22

        for i, r in enumerate(historial):
            fila = i + 3
            for col, valor in enumerate([r["ciudad"], r["maxima"], r["media"], r["minima"]], start=1):
                celda = ws2.cell(fila, col, valor)
                celda.font = Font(name="Arial", size=10)
                celda.alignment = Alignment(horizontal="center")
                if i % 2 == 0:
                    celda.fill = rp
            ws2.row_dimensions[fila].height = 18

        for col, ancho in zip("ABCD", [20, 18, 18, 18]):
            ws2.column_dimensions[col].width = ancho
        ws2.cell(len(historial) + 4, 1,
                 f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = \
            Font(name="Arial", size=9, italic=True, color="888888")

    wb.save(nombre_archivo)
    print(f"Excel guardado: {nombre_archivo}")
    return str(nombre_archivo)